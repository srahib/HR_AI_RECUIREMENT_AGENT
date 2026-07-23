import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


def export_excel(candidates, organization, job_title):

    if not os.path.exists("reports"):
        os.makedirs("reports")

    filename = "reports/candidate_report.xlsx"

    rows = []

    for i, candidate in enumerate(candidates, start=1):

        rows.append({
            "Rank": i,
            "Name": candidate["name"],
            "Email": candidate["email"],
            "Phone": candidate["phone"],
            "Age": candidate["age"],
            "Education": candidate["education"],
            "Experience": candidate["experience"],
            "Skills": ", ".join(candidate["skills"]),
            "AI Score": candidate["ai_score"],
            "Final Score": candidate["final_score"],
            "Recommendation": candidate["recommendation"]
        })

    df = pd.DataFrame(rows)

    df.to_excel(filename, index=False)

    wb = load_workbook(filename)
    ws = wb.active

    ws.insert_rows(1, amount=3)

    ws["A1"] = organization
    ws["A2"] = f"Job Title : {job_title}"
    ws["A3"] = "Top Candidates Report"

    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"].font = Font(size=12, bold=True)
    ws["A3"].font = Font(size=12, bold=True)

    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    for cell in ws[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    widths = {
        "A": 8,
        "B": 25,
        "C": 35,
        "D": 20,
        "E": 10,
        "F": 20,
        "G": 15,
        "H": 45,
        "I": 12,
        "J": 12,
        "K": 18
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(filename)

    return filename